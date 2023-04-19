#IMPORTING LIBRARIES
import datetime
import hashlib
import json
from tinyec import registry
from Crypto.Cipher import AES
import secrets
import hashlib, binascii
import pandas as pd
import numpy as np
import os
  
#CREATING BLOCKCHAIN CLASS 
class Blockchain:

    def __init__(self):
        self.chain = []
        self.create_block(proof=1, previous_hash='0')

    def create_block(self, proof, previous_hash):
        block = {'index': len(self.chain) + 1,
                 'timestamp': str(datetime.datetime.now()),
                 'proof': proof,
                 'previous_hash': previous_hash}
        self.chain.append(block)
        return block
        
    def print_previous_block(self):
        return self.chain[-1]
        
    def proof_of_work(self, previous_proof):
        new_proof = 1
        check_proof = False
          
        while check_proof is False:
            hash_operation = hashlib.sha256(
                str(new_proof**2 - previous_proof**2).encode()).hexdigest()
            if hash_operation[:4] == '0000':
                check_proof = True
            else:
                new_proof += 1
                  
        return new_proof
  
    def hash(self, block):
        encoded_block = json.dumps(block, sort_keys=True).encode()
        return hashlib.sha256(encoded_block).hexdigest()
  
    def chain_valid(self, chain):
        previous_block = chain[0]
        block_index = 1
          
        while block_index < len(chain):
            block = chain[block_index]
            if block['previous_hash'] != self.hash(previous_block):
                return False
                
            previous_proof = previous_block['proof']
            proof = block['proof']
            hash_operation = hashlib.sha256(
                str(proof**2 - previous_proof**2).encode()).hexdigest()
              
            if hash_operation[:4] != '0000':
                return False
            previous_block = block
            block_index += 1
          
        return True

#ECC ENCRYTION AND DECRYPTION WITH AES
def encryption_AES(msg, secretKey):
    aesCipher = AES.new(secretKey, AES.MODE_GCM)
    ciphertext, authTag = aesCipher.encrypt_and_digest(msg)
    return (ciphertext, aesCipher.nonce, authTag)

def decryption_AES(ciphertext, nonce, authTag, secretKey):
    aesCipher = AES.new(secretKey, AES.MODE_GCM, nonce)
    plaintext = aesCipher.decrypt_and_verify(ciphertext, authTag)
    return plaintext

def ecc_to_256_bitkey(point):
    sha = hashlib.sha256(int.to_bytes(point.x, 32, 'big'))
    sha.update(int.to_bytes(point.y, 32, 'big'))
    return sha.digest()

curve = registry.get_curve('brainpoolP256r1')


def ECC_Encrytion(msg, pubKey):
    ciphertextPrivKey = secrets.randbelow(curve.field.n)
    sharedECCKey = ciphertextPrivKey * pubKey
    secretKey = ecc_to_256_bitkey(sharedECCKey)
    ciphertext, nonce, authTag = encryption_AES(msg, secretKey)
    ciphertextPubKey = ciphertextPrivKey * curve.g
    return (ciphertext, nonce, authTag, ciphertextPubKey)

def ECC_Decrytion(storedMsg, privKey):
    (ciphertext, nonce, authTag, ciphertextPubKey) = storedMsg
    sharedECCKey = privKey * ciphertextPubKey
    secretKey = ecc_to_256_bitkey(sharedECCKey)
    plaintext = decryption_AES(ciphertext, nonce, authTag, secretKey)
    return plaintext

#-------------------------------------------------------------------------------------------------
blockchain = Blockchain()
previous_block = blockchain.print_previous_block()
previous_proof = previous_block['proof']
proof = blockchain.proof_of_work(previous_proof)
previous_hash = blockchain.hash(previous_block)
block = blockchain.create_block(proof, previous_hash)

#lOADING DATASET
df=pd.read_csv('power.csv')



#-----------------------------------------------------------------------------------------

from easygui import *
task = "Enter the Admin Login  number to be Search"
#text_query = "Enter the Query to be Search"

Key = "Enter the Key to be Search"
  
#window title
title = "Query"
task1 = enterbox(task, title)
  
# creating a integer box
#str_to_search1 = enterbox(text_query, title)

Key = passwordbox(Key, title)



if task1 == "12345":
    
    print("Reterival Cybersecurity ")
    global data1   
    data = pd.read_csv("power.csv")
    if (Key=='Password'):    
        print("Correct Key")
        import openpyxl
        import sys
        # Open the Excel workbook
        workbook = openpyxl.load_workbook("C:/Users/power.xlsx")

        # Select the worksheet 
        worksheet = workbook['Worksheet']

        # Define the search term

        search_term1 = input("ENTER THE STREET NAME:");
        search_term2 = input("ENTER THE DOORNO:");
       # search_term3 = input("Is there any applicance is bought:")
        list1=[];
        

        # Search for the search term in the worksheet
        for row in worksheet.iter_rows(values_only=True):
            if (search_term1)and(search_term2)in row:
              break;
        else:
            print("Adress not found");
            sys.exit()
              
            appliances = {
                    'Television': 100,
                    'Refrigerator': 1200,
                    'Air Conditioner': 2000,
                    'Washing Machine': 800,
                    'Electric Iron': 1000,
                    'Microwave Oven': 1200,
                    'Toaster': 800,
                    'Water Heater': 1500,
                    'Dishwasher': 1500,
                    }
    

        print(f"Is any new applicanted added newly? yes/no")
        a = str(input())
        if a == "yes":
         print(f"Enter the applicantes name ['Television','Refrigerator','Air Conditioner','Washing Machine','Electric Iron','Microwave Oven','Toaster','Water Heater','Dishwasher']")
         b = str(input())
         usage = (input(f"How many hours is the {b} used daily? "))
         total_power_daily = 23
         total_power_hourly = total_power_daily * 0.22
         total_power_weekly = total_power_daily * 7
         total_power_monthly = total_power_daily * 30
         total_power_yearly = total_power_daily * 365
         print(f"AVERAGE TOTAL POWER CONSUMED PER HOUR: {total_power_hourly} kWh")         
         print(f"AVERAGE TOTAL POWER CONSUMED PER DAY: {total_power_daily} kWh")
         print(f"AVERAGE TOTAL POWER CONSUMED PER WEEK: {total_power_weekly} kWh")
         print(f"AVERAGE TOTAL POWER CONSUMED PER MONTH: {total_power_monthly} kWh")
         print(f"AVERAGE TOTAL POWER CONSUMED PER YEAR: {total_power_yearly} kWh")
       
        else:
            list1=[];
            

            # Search for the search term in the worksheet
            for row in worksheet.iter_rows(values_only=True):
                if (search_term1)and(search_term2)in row:
                  break;
            else:
                print("Adress not found");
                sys.exit()
                  
            for i in row:
              list1.append(i);
            print("AVERAGE POWER CONSUMED PER HOUR=",list1[3]);
            print("AVERAGE POWER CONSUMED PER DAY=",list1[4]);
            print("AVERAGE POWER CONSUMED PER WEEK=",list1[5]);
            print("AVERAGE POWER CONSUMED PER MONTH=",list1[6]);
            print("AVERAGE POWER CONSUMED PER YEAR=",list1[7]);
            print("ATTACK :",list1[8]);
           
      
else:
    print("Incorrect ")
    print("=========================================================")
           
        
#------------------------------------------------------------------------------------------------------------------------
"Import Libaries "

import numpy as np
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from sklearn import metrics



print("==================================================")
print("Block chain in the Electricity   Dataset")
print(" Process - Block chain in the Electricity  Attack Detection")
print("==================================================")


##1.data slection---------------------------------------------------
#def main():
dataframe=pd.read_csv("dataset.csv")

print("---------------------------------------------")
print()
print("Data Selection")
print("Samples of our input data")
print(dataframe.head(10))
print("----------------------------------------------")
print()


 #2.pre processing--------------------------------------------------
#checking  missing values 
print("---------------------------------------------")
print()
print("Before Handling Missing Values")
print()
print(dataframe.isnull().sum())
print("----------------------------------------------")
print() 
    
print("-----------------------------------------------")
print("After handling missing values")
print()
dataframe_2=dataframe.fillna(0)
print(dataframe_2.isnull().sum())
print()
print("-----------------------------------------------")
 

#label encoding
from sklearn import preprocessing
label_encoder = preprocessing.LabelEncoder() 
print("--------------------------------------------------")
print("Before Label Handling ")
print()
print(dataframe_2.head(10))
print("--------------------------------------------------")
print()

#3.Data splitting--------------------------------------------------- 

df_train_y=dataframe_2["label"]
df_train_X=dataframe_2.iloc[:,:20]
from sklearn.preprocessing import LabelEncoder

number = LabelEncoder()

df_train_X['proto'] = number.fit_transform(df_train_X['proto'].astype(str))
df_train_X['service'] = number.fit_transform(df_train_X['service'].astype(str))
df_train_X['state'] = number.fit_transform(df_train_X['state'].astype(str))


df_train_X.head(5)
x=df_train_X
y=df_train_y
   
x_train,x_test,y_train,y_test = train_test_split(df_train_X,y,test_size = 0.20,random_state = 42)

from sklearn.ensemble import RandomForestClassifier

rf= RandomForestClassifier(n_estimators = 100)  
rf.fit(x_train, y_train)
rf_prediction = rf.predict(x_test)
Result_3=accuracy_score(y_test, rf_prediction)*100
from sklearn.metrics import confusion_matrix

print()
print("---------------------------------------------------------------------")
print("Random forest")
print()
print(metrics.classification_report(y_test,rf_prediction))
print()
print("Random forest Accuracy is:",Result_3,'%')
print()
print()
import matplotlib.pyplot as plt
import seaborn as sns




from sklearn.tree import DecisionTreeClassifier 
dt = DecisionTreeClassifier(criterion = "gini", random_state = 100,max_depth=3, min_samples_leaf=5)
dt.fit(x_train, y_train)
dt_prediction=dt.predict(x_test)
print()
print("----------------------------------------------------------------------")
print("Decision tree algorithm")
print()
Result_2=accuracy_score(y_test, dt_prediction)*100
print(metrics.classification_report(y_test,dt_prediction))
print()
print("Decision tree algorithm Accuracy is:",Result_2,'%')
print()
print("-------------------------------------------------------")
print()
import matplotlib.pyplot as plt
import seaborn as sns

